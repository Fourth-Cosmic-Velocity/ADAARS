from RPi import GPIO
from time import sleep
Import time
Import os
Import pytz
from subprocess import call
from datetime import datetime,timedelta
from picamera import PiCamera
import sounddevice as sd
from scipy.io.wavfile import write
from mpu6050 import mpu6050
from openpyxl import Workbook

GPIO.setmode(GPIO.BOARD)

GCD = 15  #Groove Coupler D0
GR = 29   # Reverse gear pin
GN =     # Neutral gear pin
G1 = 31 # First gear pin
G2 = 33 # Second gear pin
G3 = 35 # Third gear pin
G4 = 37 # Fourth gear pin
G5 = 40 # Fifth gear pin
DREC = 22 # Impact Switch
RECL = 07 # Rotary Encoder Clock
REDA = 11 # Rotary Encoder Data
%REDS = 13 # Rotary Encoder Switch
SBS = 16 # Seat Belt switch
USTRIG = 32 # USD Trigger pin
USECHO = 36 # USD Echo pin

GPIO.setup(GCD,GPIO.IN)
GPIO.setup(GR, GPIO.IN, pull_up_down=GPIO.PUD_DOWN)
GPIO.setup(G1, GPIO.IN, pull_up_down=GPIO.PUD_DOWN)
GPIO.setup(G2, GPIO.IN, pull_up_down=GPIO.PUD_DOWN)
GPIO.setup(G3, GPIO.IN, pull_up_down=GPIO.PUD_DOWN)
GPIO.setup(G4, GPIO.IN, pull_up_down=GPIO.PUD_DOWN)
GPIO.setup(G5, GPIO.IN, pull_up_down=GPIO.PUD_DOWN)
GPIO.setup(IMPCT, GPIO.IN, pull_up_down=GPIO.PUD_DOWN)
GPIO.setup(RECL,GPIO.IN)
GPIO.setup(REDA,GPIO.IN)
GPIO.setup(REDS,GPIO.IN)
GPIO.setup(SBS, GPIO.IN, pull_up_down=GPIO.PUD_DOWN)
GPIO.setup(USTRIG,GPIO.OUT)
GPIO.setup(USECHO,GPIO.IN)


def  Impact():
    while True:
    IMP=GPIO.input(IMPCT)
wb = Workbook()
ws = wb.active
ws['A1'] = 'Date'
ws['B1'] = 'Time'
ws['C1'] = 'X'
ws['D1'] = 'Y'
ws['E1'] = 'Z'
ws['F1'] = 'Pitch'
ws['G1'] = 'Roll'
ws['H1'] = 'Yaw'
ws['I1'] = 'Temperature'
ws['J1'] = 'SWA'
ws['K1'] = 'BPD(cm)'
ws['L1'] = 'Speed(Hz)'
ws['M1'] = 'Gear'
ws['N1'] = 'Impact'
ws[‘O1’]=’Seat Belt’
counter=0
i=2

    while(IMP==1):
        for k in range(0,300):
            now=datetime.now()
                ws['A%s'%i]=now.strftime("%Y:%m:%d")
ws['B%s'%i]=now.strftime("%H:%M:%S")    ws['C%s'%i],ws['D%s'%i],ws['E%s'%i],ws['F%s'%i],ws['G%s'%i],ws['H%s'%i],ws['I%s'%i]=IMU()
ws['J%s'%i]=Steering_Wheel_Angle()
ws['K%s'%i]=Brake_Pedal_Height()
ws['L%s'%i]=Vehicle_Speed()
ws['M%s'%i]=Gear_Position
ws['N%s'%i]=IMP
ws['O%s'%i]=Seatbelt()
                i=i+1
                time.sleep(1)
            wb.save("sample.xlsx")

    while (IMP==0):
            now=datetime.now()
            ws['A%s'%i]=now.strftime("%Y:%m:%d")
ws['B%s'%i]=now.strftime("%H:%M:%S")
ws['C%s'%i],ws['D%s'%i],ws['E%s'%i],ws['F%s'%i],ws['G%s'%i],ws['H%s'%i],ws['I%s'%i]=IMU()
ws['J%s'%i]=Steering_Wheel_Angle()
ws['K%s'%i]=Brake_Pedal_Height()
ws['L%s'%i]=Vehicle_Speed()
ws['M%s'%i]=Gear_Position
ws['N%s'%i]=IMP
ws['O%s'%i]=Seatbelt()
           counter=counter+1
         i=i+1
            time.sleep(1)
            if counter>600:
                    ws.move_range("A3:N602", rows=-1)
wb.save("sample.xlsx")
“ ” ”
        
def IMU():
    
mpu = mpu6050(0x68)
        print("Temp : "+str(mpu.get_temp()))
       print()

        accel_data = mpu.get_accel_data()
        print("Acc X : "+str(accel_data['x']))
        print("Acc Y : "+str(accel_data['y']))
        print("Acc Z : "+str(accel_data['z']))
        print()

        gyro_data = mpu.get_gyro_data()
        print("Gyro X : "+str(gyro_data['x']))
        print("Gyro Y : "+str(gyro_data['y']))
        print("Gyro Z : "+str(gyro_data['z']))
print()    return(str(accel_data['x']),str(accel_data['y']),str(accel_data['z']),str(gyro_data['x']),str(gyro_data['y']),str(gyro_data['z']),str(mpu.get_temp()))
“ ” ”

def Vehicle_Speed():
    GCST= int(round(time.time() * 1000))
GCET = GCST + 1000
WROT=0
OLD_WROT=0
while time.time()*1000 < GCET:
    if GPIO.input(GCD):
        WROT=WROT+1
        while GPIO.input(GCD)
STEPS=WROT-OLD_WROT
OLD_WROT=WROT
    RPM=STEPS*60/20
    return RPM

def Brake_Pedal_Height():
print("distance measurement in progress")
GPIO.setup(USTRIG,GPIO.OUT)
GPIO.setup(ECHO,GPIO.IN)
GPIO.output(USTRIG,False) 
print("waiting for sensor to settle")
time.sleep(0.2)
GPIO.output(USTRIG,True)
time.sleep(0.00001) 
GPIO.output(USTRIG,False) 
while GPIO.input(ECHO)==0: 
pulse_start=time.time()
while GPIO.input(ECHO)==1: 
pulse_end=time.time() 
pulse_duration=pulse_end-pulse_start
distance=pulse_duration*17150 
distance=round(distance,2) 
print("distance:",distance,"cm") 

def Steering_Wheel_Angle():
counter = 0
clkLastState = GPIO.input(RECL)

try: 
clkState = GPIO.input(RECL)
             dtState = GPIO.input(REDA)
              if clkState != clkLastState:
                        if dtState != clkState:
                                counter += 1
                        else:
                                counter -= 1
                        print (counter*9)
                clkLastState = clkState
                time.sleep(0.01)
finally:
        GPIO.cleanup()
        return counter*9

def Seatbelt():
SB=GPIO.input(SBS)
               print(SB)
    return SB


def Gear_Position():
    if (GPIO.input(GR)):
        GP= “Reverse Gear”
    elif (GPIO.input(G1)):
        GP= “First Gear”
    elif (GPIO.input(G2)):
        GP= “Second Gear”
    elif (GPIO.input(G3)):
        GP= “Third Gear”
    elif (GPIO.input(G4)):
        GP= “Fourth Gear”
    elif (GPIO.input(G5)):
        GP= “Fifth Gear”
    print(GP)
    return GP

def Video():
    camera = PiCamera()
i=0
while(true):
           now = datetime.now(pytz.timezone(‘Asia/Calcutta’))
            current_time = now.strftime("%H:%M:%S")
            ten_minute=timedelta(minutes=10)
            camera.start_preview()
            camera.start_recording('/home/pi/Desktop/%s.h264' %current_time)
           camera.wait_recording(60)
            camera.stop_recording()
            camera.stop_preview()
            i=i+1
            if(i>10):
                    old_time=(now-ten_minute).strftime("%H:%M:%S")
                    os.remove('/home/pi/Desktop/%s.h264' %old_time)
def Audio():
i=0
while(true):
            now = datetime.now(pytz.timezone('Asia/Calcutta'))
            current_time = now.strftime("%H-%M-%S")
            ten_minute=timedelta(minutes=10)
            duration = 60  
            fs = 48000 #bitrate
            myrecording = sd.rec(int(duration * fs), samplerate=fs, channels=1)
            sd.wait()
            write("%s.wav" %current_time,fs,myrecording)
             i=i+1
            if(i>10):
                old_time=(now-ten_minute).strftime("%H-%M-%S")
                os.remove('/home/pi/Desktop/%s.wav' %old_time)

Impact = threading.Thread(target=Impact)
IMU = threading.Thread(target=IMU)
Vehicle_Speed = threading.Thread(target=Vehicle_Speed)
Brake_Pedal_Height = threading.Thread(target=Brake_Pedal_Height)
Steering_Wheel_Angle = threading.Thread(target=Steering_Wheel_Angle)
Seatbelt = threading.Thread(target=Seatbelt)
Gear_Position = threading.Thread(target=Gear_Position)
Video = threading.Thread(target=Video)
Audio = threading.Thread(target=Audio)
